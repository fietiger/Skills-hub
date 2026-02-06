import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment, Font
from copy import copy
import datetime
import os
import re

def generate_expense_list():
    # 1. 定义文件路径 (使用相对路径或从 skill 资源目录读取)
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    template_path = os.path.join(base_dir, 'assets', 'expense_template.xlsx')
    
    # 输入文件默认在当前工作目录
    train_path = os.path.abspath('火车票汇总信息表.xlsx')
    didi_path = os.path.abspath('滴滴行程明细信息汇总表.xlsx')
    accommodation_path = os.path.abspath('住宿发票信息汇总表.xlsx')
    meal_path = os.path.abspath('餐费发票信息汇总表.xlsx')
    output_path = os.path.abspath('费用清单.xlsx')

    def clean_price(val):
        if val is None: return 0.0
        # 清理 '=' '¥' ',' 等符号
        s = str(val).replace('=', '').replace('¥', '').replace(',', '').strip()
        try:
            return float(s)
        except:
            return 0.0

    # 2. 整理汇总数据
    consolidated_data = []

    # 处理火车票（如果文件存在）
    if os.path.exists(train_path):
        df_train = pd.read_excel(train_path)
        for _, row in df_train.iterrows():
            price = clean_price(row['price'])
            dep = str(row['departure_station'])
            arr = str(row['arrival_station'])
            date_str = str(row['date'])[:10]  # 取 YYYY-MM-DD 部分
            if '/' in date_str:
                # 如果日期格式是 YYYY/MM/DD，转换为 YYYY-MM-DD
                date_str = date_str.replace('/', '-')
            
            consolidated_data.append({
                '日期': date_str,
                '事由': f"出差交通({dep}-{arr})",
                '项目名称': '公共项目',
                '类别': '长途交通费',
                '金额': price,
                '备注': f"车次: {row['train_number']}"
            })
        print(f"已处理火车票数据: {len(df_train)} 条记录")
    else:
        print(f"警告: 未找到火车票文件 {train_path}，将跳过火车票处理")

    # 处理滴滴行程（如果文件存在）
    if os.path.exists(didi_path):
        df_didi = pd.read_excel(didi_path)
        for _, row in df_didi.iterrows():
            time_str = str(row['上车时间'])
            try:
                # 补全年份为 2026
                dt = datetime.datetime.strptime(f"2026-{time_str}", '%Y-%m-%d %H:%M')
                date_str = dt.strftime('%Y-%m-%d')
            except:
                date_str = time_str
            
            consolidated_data.append({
                '日期': date_str,
                '事由': '市内交通',
                '项目名称': '公共项目',
                '类别': '市内交通费',
                '金额': row['金额[元]'],
                '备注': ''
            })
        print(f"已处理滴滴行程数据: {len(df_didi)} 条记录")
    else:
        print(f"警告: 未找到滴滴行程文件 {didi_path}，将跳过滴滴行程处理")

    # 处理住宿发票（如果文件存在）
    if os.path.exists(accommodation_path):
        df_accommodation = pd.read_excel(accommodation_path)
        for _, row in df_accommodation.iterrows():
            # 从开票日期提取日期部分 - 处理不同格式的日期
            date_val = row['开票日期']
            date_str = str(date_val)
            
            # 如果是 YYYYMMDD 格式，转换为 YYYY-MM-DD
            if len(date_str) == 8 and date_str.isdigit():
                date_str = f"{date_str[:4]}-{date_str[4:6]}-{date_str[6:8]}"
            elif '/' in date_str:
                # 如果日期格式是 YYYY/MM/DD，转换为 YYYY-MM-DD
                date_str = date_str.replace('/', '-')
            elif '年' in date_str and '月' in date_str and '日' in date_str:
                # 处理 "YYYY年MM月DD日" 格式
                match = re.search(r'(\d{4})年(\d{1,2})月(\d{1,2})日', date_str)
                if match:
                    year, month, day = match.groups()
                    date_str = f"{year}-{int(month):02d}-{int(day):02d}"
            
            consolidated_data.append({
                '日期': date_str,
                '事由': '住宿费',
                '项目名称': '公共项目',
                '类别': '住宿费',
                '金额': clean_price(row['金额']),
                '备注': f"酒店: {row['销售方名称']}"
            })
        print(f"已处理住宿发票数据: {len(df_accommodation)} 条记录")
    else:
        print(f"警告: 未找到住宿发票文件 {accommodation_path}，将跳过住宿发票处理")

    # 处理餐费发票（如果文件存在）
    if os.path.exists(meal_path):
        df_meal = pd.read_excel(meal_path)
        for _, row in df_meal.iterrows():
            # 从开票日期提取日期部分 - 处理不同格式的日期
            date_val = row['开票日期']
            date_str = str(date_val)
            
            # 如果是 YYYYMMDD 格式，转换为 YYYY-MM-DD
            if len(date_str) == 8 and date_str.isdigit():
                date_str = f"{date_str[:4]}-{date_str[4:6]}-{date_str[6:8]}"
            elif '/' in date_str:
                # 如果日期格式是 YYYY/MM/DD，转换为 YYYY-MM-DD
                date_str = date_str.replace('/', '-')
            elif '年' in date_str and '月' in date_str and '日' in date_str:
                # 处理 "YYYY年MM月DD日" 格式
                match = re.search(r'(\d{4})年(\d{1,2})月(\d{1,2})日', date_str)
                if match:
                    year, month, day = match.groups()
                    date_str = f"{year}-{int(month):02d}-{int(day):02d}"
            
            consolidated_data.append({
                '日期': date_str,
                '事由': '餐费',
                '项目名称': '公共项目',
                '类别': '餐费',
                '金额': clean_price(row['金额']),
                '备注': f"商家: {row['销售方名称']}"
            })
        print(f"已处理餐费发票数据: {len(df_meal)} 条记录")
    else:
        print(f"警告: 未找到餐费发票文件 {meal_path}，将跳过餐费发票处理")

    # 按日期排序
    df_result = pd.DataFrame(consolidated_data).sort_values(by='日期')

    # 4. 写入模板并设置格式
    wb = load_workbook(template_path)
    ws = wb.active

    # 提取第 4 行或第 3 行作为样式模板
    sample_row = ws[4] if ws.max_row >= 4 else ws[3]
    styles = []
    for cell in sample_row:
        styles.append({
            'font': copy(cell.font),
            'border': copy(cell.border),
            'fill': copy(cell.fill),
            'alignment': copy(cell.alignment)
        })

    # 清除原有数据 (从第 4 行开始)
    while ws.max_row >= 4:
        ws.delete_rows(4)

    # 货币格式 (¥ 符号 + 2位小数)
    rmb_format = '¥#,##0.00'

    # 填充新数据
    for idx, (_, row) in enumerate(df_result.iterrows()):
        target_row = 4 + idx
        for col_idx, value in enumerate([row['日期'], row['事由'], row['项目名称'], 
                                         row['类别'], row['金额'], row['备注']], start=1):
            cell = ws.cell(row=target_row, column=col_idx, value=value)
            if col_idx <= len(styles):
                style = styles[col_idx-1]
                cell.font = style['font']
                cell.border = style['border']
                cell.fill = style['fill']
                cell.alignment = style['alignment']

        # 设置货币格式
        if isinstance(row['金额'], (int, float)):
            ws.cell(row=target_row, column=5).number_format = rmb_format

    # 添加合计行
    total_row = 4 + len(df_result)
    ws.cell(row=total_row, column=2, value="合计").font = Font(bold=True)
    ws.cell(row=total_row, column=5, value=f"=SUM(E4:E{total_row-1})").number_format = rmb_format

    # 5. 特殊处理 A2:F2 合并单元格边框 (移除整个范围的左右外边框)
    # 对于合并单元格，需要处理范围边界上的所有单元格
    for col in range(1, 7): # A to F
        cell = ws.cell(row=2, column=col)
        current_border = cell.border
        left_side = Side(style=None) if col == 1 else current_border.left
        right_side = Side(style=None) if col == 6 else current_border.right
        
        cell.border = Border(
            left=left_side,
            right=right_side,
            top=current_border.top,
            bottom=current_border.bottom
        )


    # 6. 保存结果
    wb.save(output_path)
    print(f"成功生成最终费用清单: {output_path}")

if __name__ == "__main__":
    generate_expense_list()
