import pandas as pd
from openpyxl import load_workbook

def fill_expense_reimbursement():
    """
    从费用清单.xlsx自动填写费用报销单.xlsx的E5-E7单元格
    """
    try:
        # 读取费用清单来获取数据
        expense_df = pd.read_excel('费用清单.xlsx')
        
        # 分析费用清单，计算各类费用总额
        transportation_total = 0  # 交通费用（火车票+滴滴票）
        accommodation_total = 0   # 住宿费用
        other_total = 0          # 其他费用
        
        # 遍历数据行（跳过前两行标题）
        for idx in range(2, len(expense_df)-1):  # -1是为了排除最后一行的"合计"
            reason = str(expense_df.iloc[idx, 1])  # 事由列
            amount = expense_df.iloc[idx, 4]       # 金额列（E列，索引为4）
            
            if pd.isna(amount):  # 如果金额为空，跳过
                continue
                
            if '交通' in reason:
                transportation_total += amount
            elif '住宿' in reason:
                accommodation_total += amount
            else:
                other_total += amount
        
        print(f"费用分类统计:")
        print(f"交通费用（火车票+滴滴票）: {transportation_total}")
        print(f"住宿费用: {accommodation_total}")
        print(f"其他费用: {other_total}")
        
        # 使用openpyxl直接操作Excel文件
        # 从技能资源目录读取费用报销单模板
        template_path = '.trae/skills/expense-reimbursement-filler/assets/费用报销单.xlsx'
        workbook = load_workbook(template_path)
        worksheet = workbook.active
        
        # 根据要求，直接更新E5、E6、E7单元格
        # E列是第5列，所以：
        worksheet.cell(row=5, column=5, value=transportation_total)  # E5 = 交通费
        worksheet.cell(row=6, column=5, value=accommodation_total)   # E6 = 住宿费
        worksheet.cell(row=7, column=5, value=other_total)           # E7 = 其他费用
        
        # 计算并更新合计
        total_amount = transportation_total + accommodation_total + other_total
        worksheet.cell(row=8, column=5, value=total_amount)          # E8 = 合计
        
        print(f"更新的单元格:")
        print(f"E5 (交通费): {transportation_total}")
        print(f"E6 (住宿费): {accommodation_total}")
        print(f"E7 (其他费用): {other_total}")
        print(f"E8 (合计): {total_amount}")
        
        # 保存文件
        workbook.save('费用报销单_已填写.xlsx')
        print(f"已将填写好的费用报销单保存为 '费用报销单_已填写.xlsx'")
        
        # 验证保存的文件
        workbook_saved = load_workbook('费用报销单_已填写.xlsx')
        worksheet_saved = workbook_saved.active
        
        print(f"验证保存的文件中关键单元格:")
        print(f"E5: {worksheet_saved.cell(row=5, column=5).value}")
        print(f"E6: {worksheet_saved.cell(row=6, column=5).value}")
        print(f"E7: {worksheet_saved.cell(row=7, column=5).value}")
        print(f"E8: {worksheet_saved.cell(row=8, column=5).value}")
        
        return True
        
    except Exception as e:
        print(f"处理过程中出现错误: {e}")
        return False

if __name__ == "__main__":
    fill_expense_reimbursement()