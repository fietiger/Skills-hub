from openpyxl import load_workbook
import sys
import os

def update_reimbursement_info(file_path, page_count, reimbursename):
    """
    更新报销单中的页数和报销人信息
    """
    # 检查文件是否存在
    if not os.path.exists(file_path):
        print(f"错误: 文件 {file_path} 不存在")
        return False
    
    try:
        # 读取报销单
        wb = load_workbook(file_path)
        ws = wb.active

        # 更新J3单元格的页数信息 (单据及附件共X页)
        ws['J3'] = f'单据及附件共{page_count}页'
        print(f'J3单元格: {ws["J3"].value}')

        # 更新报销人信息到I9单元格
        ws['I9'] = reimbursename
        print(f'I9单元格: {ws["I9"].value}')

        # 保存文件
        wb.save(file_path)
        print(f'报销单已更新完成: {file_path}')
        return True
        
    except Exception as e:
        print(f"更新报销单时发生错误: {str(e)}")
        return False

if __name__ == "__main__":
    if len(sys.argv) < 4:
        print("用法: python update_reimbursement_info.py <excel文件路径> <页数> <报销人姓名>")
        print("示例: python update_reimbursement_info.py 费用报销单_已填写.xlsx 11 程文清")
        sys.exit(1)
    
    file_path = sys.argv[1]
    page_count = int(sys.argv[2])
    reimbursename = sys.argv[3]
    
    success = update_reimbursement_info(file_path, page_count, reimbursename)
    if not success:
        sys.exit(1)